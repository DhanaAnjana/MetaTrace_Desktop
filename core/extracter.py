import os
import sys
from pathlib import Path
from datetime import datetime


FILE_SIGNATURES = {
    b'\xFF\xD8\xFF': ['.jpg', '.jpeg'],
    b'\x89PNG': '.png',
    b'GIF8': '.gif',
    b'\x42\x4D': '.bmp',
    b'%PDF': '.pdf',
    b'PK\x03\x04': ['.docx', '.xlsx', '.pptx', '.zip'],
    b'\xFF\xFB': '.mp3',
    b'ID3': '.mp3',
    b'\x00\x00\x00\x18ftypmp4': '.mp4',
    b'{': '.json',
    b'From:': '.eml',
}


def detect_office_file_type(file_path):
    try:
        import zipfile
        
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            namelist = zip_ref.namelist()
            
            if any('ppt/' in name for name in namelist):
                return '.pptx'
            elif any('xl/' in name for name in namelist):
                return '.xlsx'
            elif any('word/' in name for name in namelist):
                return '.docx'
            else:
                return '.zip'
    except:
        return '.zip'


def detect_file_type(file_path):
    try:
        with open(file_path, 'rb') as f:
            header = f.read(32)
        
        for signature, ext_list in FILE_SIGNATURES.items():
            if header.startswith(signature):
                exts = ext_list if isinstance(ext_list, list) else [ext_list]
                
                if signature == b'PK\x03\x04':
                    detected = detect_office_file_type(file_path)
                    return detected, True
                
                detected_ext = exts[0]
                return detected_ext, True
        
        return Path(file_path).suffix, False
    except:
        return Path(file_path).suffix, False


def validate_file_signature(file_path):
    actual_extension = Path(file_path).suffix.lower()
    detected_extension, is_real = detect_file_type(file_path)
    
    if not is_real:
        return {
            "valid": True,
            "warning": None,
            "detected_type": detected_extension,
            "claimed_type": actual_extension
        }
    
    if actual_extension in ['.jpg', '.jpeg']:
        if detected_extension in ['.jpg', '.jpeg']:
            return {
                "valid": True,
                "warning": None,
                "detected_type": detected_extension,
                "claimed_type": actual_extension
            }
    
    if detected_extension != actual_extension:
        return {
            "valid": False,
            "warning": f"File signature mismatch! Claimed: {actual_extension}, Actual: {detected_extension}",
            "detected_type": detected_extension,
            "claimed_type": actual_extension
        }
    
    return {
        "valid": True,
        "warning": None,
        "detected_type": detected_extension,
        "claimed_type": actual_extension
    }


def extract_basic_info(file_path):
    info = {}
    
    file_path = Path(file_path)
    if not file_path.exists():
        return {"error": "File not found"}
    
    stat = file_path.stat()
    info["file_name"] = file_path.name
    info["file_path"] = str(file_path)
    info["file_size"] = stat.st_size
    
    created_dt = datetime.fromtimestamp(stat.st_ctime)
    modified_dt = datetime.fromtimestamp(stat.st_mtime)
    
    info["created_time"] = created_dt.strftime("%Y-%m-%d %H:%M:%S")
    info["modified_time"] = modified_dt.strftime("%Y-%m-%d %H:%M:%S")
    info["file_extension"] = file_path.suffix
    
    sig_info = validate_file_signature(str(file_path))
    info["signature_valid"] = sig_info["valid"]
    if sig_info["warning"]:
        info["signature_warning"] = sig_info["warning"]
    
    return info


def detect_anomalies(file_path):
    anomalies = []
    
    file_path = Path(file_path)
    stat = file_path.stat()
    
    created = float(stat.st_ctime)
    modified = float(stat.st_mtime)
    
    if created > modified:
        anomalies.append({
            "type": "temporal_anomaly",
            "message": "Created date is after modified date",
            "severity": "high"
        })
    
    if stat.st_size == 0:
        anomalies.append({
            "type": "empty_file",
            "message": "File is empty",
            "severity": "medium"
        })
    
    sig_info = validate_file_signature(str(file_path))
    if not sig_info["valid"]:
        anomalies.append({
            "type": "signature_mismatch",
            "message": sig_info["warning"],
            "severity": "high"
        })
    
    return anomalies


def extract_image_metadata(file_path):
    info = extract_basic_info(file_path)
    from PIL import Image
    from PIL.ExifTags import TAGS
    
    image = Image.open(file_path)
    info["image_format"] = image.format
    info["image_size"] = image.size
    info["image_mode"] = image.mode
    
    exif_data = image._getexif()
    if exif_data:
        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, tag_id)
            info[f"exif_{tag_name}"] = str(value)[:100]
    return info


def extract_pdf_metadata(file_path):
    info = extract_basic_info(file_path)
    
    from PyPDF2 import PdfReader
    
    pdf = PdfReader(file_path)
    info["pdf_pages"] = len(pdf.pages)
    
    info["pdf_author"] = ""
    info["pdf_creation_date"] = ""
    info["pdf_mod_date"] = ""
    info["pdf_producer"] = ""
    info["pdf_title"] = ""
    info["pdf_subject"] = ""
    
    if pdf.metadata:
        info["pdf_author"] = pdf.metadata.get("/Author") or ""
        info["pdf_creation_date"] = pdf.metadata.get("/CreationDate") or ""
        info["pdf_mod_date"] = pdf.metadata.get("/ModDate") or ""
        info["pdf_producer"] = pdf.metadata.get("/Producer") or ""
        info["pdf_title"] = pdf.metadata.get("/Title") or ""
        info["pdf_subject"] = pdf.metadata.get("/Subject") or ""
    
    return info


def extract_office_metadata(file_path):
    info = extract_basic_info(file_path)
    extension = Path(file_path).suffix.lower()
    
    if extension == ".docx":
        from docx import Document
        doc = Document(file_path)
        info["word_paragraphs"] = len(doc.paragraphs)
        core_props = doc.core_properties
        info["word_author"] = core_props.author or ""
        info["word_title"] = core_props.title or ""
        info["word_subject"] = core_props.subject or ""
        info["word_created"] = str(core_props.created) if core_props.created else ""
        info["word_modified"] = str(core_props.modified) if core_props.modified else ""
        info["word_last_modified_by"] = core_props.last_modified_by or ""
        
    elif extension == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        info["excel_sheets"] = len(wb.sheetnames)
        info["excel_sheet_names"] = wb.sheetnames
        
    elif extension == ".pptx":
        from pptx import Presentation
        prs = Presentation(file_path)
        info["pptx_slides"] = len(prs.slides)
    
    return info


def extract_file_metadata(file_path):
    path = Path(file_path)
    extension = path.suffix.lower()
    
    if extension in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
        return extract_image_metadata(file_path)
    elif extension == ".pdf":
        return extract_pdf_metadata(file_path)
    elif extension in [".docx", ".xlsx", ".pptx"]:
        return extract_office_metadata(file_path)
    else:
        return extract_basic_info(file_path)


def extract_multiple_files(file_paths):
    results = []
    for file_path in file_paths:
        metadata = extract_file_metadata(file_path)
        from core.hash_utils import calculate_all_hashes
        hashes = calculate_all_hashes(file_path)
        anomalies = detect_anomalies(file_path)
        
        result = {
            "file_path": file_path,
            "metadata": metadata,
            "hashes": hashes,
            "anomalies": anomalies
        }
        results.append(result)
    
    return results


def scan_folder(folder_path, recursive=False):
    results = []
    folder = Path(folder_path)
    
    if not folder.is_dir():
        return {"error": "Not a valid folder"}
    
    pattern = "**/*" if recursive else "*"
    
    for file_path in folder.glob(pattern):
        if file_path.is_file():
            metadata = extract_file_metadata(str(file_path))
            from core.hash_utils import calculate_all_hashes
            hashes = calculate_all_hashes(str(file_path))
            anomalies = detect_anomalies(str(file_path))
            
            result = {
                "file_path": str(file_path),
                "metadata": metadata,
                "hashes": hashes,
                "anomalies": anomalies
            }
            results.append(result)
    
    return results